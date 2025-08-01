from fastapi import FastAPI, UploadFile, Form, WebSocket, BackgroundTasks
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import os, shutil, json, subprocess, asyncio, tempfile
from uuid import uuid4
from openpyxl import load_workbook
from openpyxl.worksheet.pagebreak import Break, RowBreak
from openpyxl.utils import column_index_from_string
from PIL import Image
from io import BytesIO

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

OUTPUT_BASE = "outputs"
os.makedirs(OUTPUT_BASE, exist_ok=True)

progress_map = {}  # job_id -> { step, percent, status }


def update_progress(job_id, step, percent):
    progress_map[job_id] = {"step": step, "percent": percent}


def cleanup_output_directory():
    """Clean up the output directory by removing old files and temp directories"""
    try:
        import time
        current_time = time.time()
        cleaned_count = 0
        
        print(f"Starting cleanup of output directory: {OUTPUT_BASE}")
        
        if not os.path.exists(OUTPUT_BASE):
            print(f"Output directory {OUTPUT_BASE} does not exist")
            return
            
        for item in os.listdir(OUTPUT_BASE):
            item_path = os.path.join(OUTPUT_BASE, item)
            
            # Remove old PNG files (older than 10 minutes)
            if item.endswith('.png'):
                age_seconds = current_time - os.path.getctime(item_path)
                if age_seconds > 0:  # 10 minutes
                    print(f"Removing old PNG: {item} (age: {age_seconds:.1f} seconds)")
                    os.remove(item_path)
                    cleaned_count += 1
            
            # Remove temp directories (older than 5 minutes)
            elif os.path.isdir(item_path) and item.startswith('temp_'):
                age_seconds = current_time - os.path.getctime(item_path)
                if age_seconds > 0:  # 5 minutes
                    print(f"Removing temp directory: {item} (age: {age_seconds:.1f} seconds)")
                    shutil.rmtree(item_path, ignore_errors=True)
                    cleaned_count += 1
        
        print(f"Cleanup completed. Removed {cleaned_count} items.")
        
    except Exception as e:
        print(f"Error cleaning up output directory: {e}")
        import traceback
        traceback.print_exc()


def delete_rows_and_columns(ws, ranges, cols_to_delete):
    original_merges = list(ws.merged_cells.ranges)
    merge_text_map = {}
    for merge in original_merges:
        val = ws.cell(row=merge.min_row, column=merge.min_col).value
        if val:
            merge_text_map[(merge.min_row, merge.min_col)] = val

    col_widths = {chr(64 + col): ws.column_dimensions[chr(64 + col)].width or 8.43 for col in range(1, ws.max_column + 1)}
    row_heights = {row: dim.height for row, dim in ws.row_dimensions.items() if dim.height}

    for col in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col)

    adjusted_widths = {}
    for col, width in col_widths.items():
        idx = ord(col) - 64 - len(cols_to_delete)
        if idx > 0:
            adjusted_widths[chr(64 + idx)] = width

    row_map, new_row = {}, 1
    for start, end in ranges:
        for old in range(start, end + 1):
            row_map[old] = new_row
            new_row += 1

    for row in range(ws.max_row, 0, -1):
        if row not in row_map:
            ws.delete_rows(row)

    ws.merged_cells.ranges = []
    for merge in original_merges:
        min_r, min_c, max_r, max_c = merge.min_row, merge.min_col, merge.max_row, merge.max_col
        if min_r in row_map and max_r in row_map:
            ws.merge_cells(
                start_row=row_map[min_r],
                start_column=max(1, min_c - len(cols_to_delete)),
                end_row=row_map[max_r],
                end_column=max(1, max_c - len(cols_to_delete)),
            )
            if (merge.min_row, merge.min_col) in merge_text_map:
                ws.cell(row=row_map[min_r], column=max(1, min_c - len(cols_to_delete))).value = merge_text_map[(merge.min_row, merge.min_col)]

    for col, width in adjusted_widths.items():
        ws.column_dimensions[col].width = width
    for old, new in row_map.items():
        if old in row_heights:
            ws.row_dimensions[new].height = row_heights[old]


def insert_page_breaks(ws, ranges):
    ws.row_breaks = RowBreak()
    pos = 0
    for i, (start, end) in enumerate(ranges):
        pos += (end - start + 1)
        if i < len(ranges) - 1:
            ws.row_breaks.append(Break(id=pos))


def process_sheet(ws, ranges, columns_range):
    # Parse "C:J" into column indices
    start_col, end_col = columns_range.split(":")
    start_idx = column_index_from_string(start_col)
    end_idx = column_index_from_string(end_col)

    # Determine columns to delete
    cols_to_delete = list(range(end_idx + 1, ws.max_column + 1)) + list(range(1, start_idx))

    delete_rows_and_columns(ws, ranges, cols_to_delete)
    insert_page_breaks(ws, ranges)
    ws.print_area = f"A1:{ws.dimensions.split(':')[1]}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def crop_image(input_path, output_path, padding=5):
    img = Image.open(input_path).convert("RGB")
    pixels, bg = img.load(), (255, 255, 255)
    left, top, right, bottom = img.width, img.height, 0, 0
    for y in range(img.height):
        for x in range(img.width):
            if pixels[x, y] != bg:
                left, top, right, bottom = min(left, x), min(top, y), max(right, x), max(bottom, y)
    if left < right and top < bottom:
        img.crop(
            (max(left - padding, 0), max(top - padding, 0), min(right + padding, img.width), min(bottom + padding, img.height))
        ).save(output_path)


async def run_processing_job(job_id, input_path, parsed_config, session_dir):
    update_progress(job_id, "Processing started", 5)
    wb = load_workbook(input_path)

    sheets_data = parsed_config["sheets"]

    # Keep only selected sheets
    selected_indices = [s["index"] for s in sheets_data]
    all_sheetnames = wb.sheetnames

    for idx, name in enumerate(all_sheetnames, start=1):
        if idx not in selected_indices:
            ws_to_remove = wb[name]
            wb.remove(ws_to_remove)

    for i, sheet_conf in enumerate(sheets_data):
        ws = wb.worksheets[sheet_conf["index"] - 1]
        ranges = [(a["start"], a["end"]) for a in sheet_conf["assets"]]
        columns_range = sheet_conf["columns"]
        process_sheet(ws, ranges, columns_range)
        update_progress(job_id, f"Processing sheet {sheet_conf['index']}", 10 + int((i + 1) / len(sheets_data) * 20))
        await asyncio.sleep(0.1)

    modified_excel = os.path.join(session_dir, "processed.xlsx")
    wb.save(modified_excel)

    update_progress(job_id, "Converting to PDF", 40)
    subprocess.run(f'soffice --headless --convert-to pdf "{modified_excel}" --outdir "{session_dir}"', shell=True, check=True)
    await asyncio.sleep(0.1)

    update_progress(job_id, "Converting to PNG", 60)
    subprocess.run(f'pdftoppm -png -rx 300 -ry 300 "{os.path.join(session_dir,"processed.pdf")}" "{os.path.join(session_dir,"page")}"', shell=True, check=True)

    png_urls = []
    all_files = sorted([f for f in os.listdir(session_dir) if f.endswith(".png")])
    for idx, f in enumerate(all_files):
        # Create unique filename for output folder
        unique_filename = f"{job_id}_{f}"
        output_path = os.path.join(OUTPUT_BASE, unique_filename)
        
        # Crop and save directly to outputs folder
        crop_image(os.path.join(session_dir, f), output_path)
        update_progress(job_id, "Cropping images", 60 + int((idx + 1) / len(all_files) * 35))
        await asyncio.sleep(0.1)
        png_urls.append(f"/files/{unique_filename}")

    # Clean up the temporary session directory
    try:
        shutil.rmtree(session_dir, ignore_errors=True)
        print(f"Cleaned up temporary session directory: {session_dir}")
    except Exception as e:
        print(f"Error cleaning up session directory: {e}")

    update_progress(job_id, "Done", 100)
    progress_map[job_id]["images"] = png_urls


# Serve static files
app.mount("/files", StaticFiles(directory="outputs"), name="files")


@app.websocket("/progress/{job_id}")
async def progress_ws(websocket: WebSocket, job_id: str):
    await websocket.accept()
    while True:
        if job_id in progress_map:
            await websocket.send_json(progress_map[job_id])
            if progress_map[job_id]["percent"] >= 100:
                break
        await asyncio.sleep(0.5)
    await websocket.close()


@app.post("/process-excel/")
async def process_excel(
    file: UploadFile,
    config: str = Form(...),  # Full JSON config for sheets/assets/columns
    background_tasks: BackgroundTasks = None
):
    # Clean up old output directories before processing
    cleanup_output_directory()
    
    job_id = str(uuid4())
    update_progress(job_id, "File upload", 0)

    parsed_config = json.loads(config)
    session_dir = os.path.join(OUTPUT_BASE, job_id)
    os.makedirs(session_dir, exist_ok=True)

    input_path = os.path.join(session_dir, file.filename)
    with open(input_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    update_progress(job_id, "Uploaded", 5)

    # Run the processing job in the background
    background_tasks.add_task(run_processing_job, job_id, input_path, parsed_config, session_dir)

    return JSONResponse({"job_id": job_id})


@app.post("/add-assets/")
async def add_assets(
    file: UploadFile,
    config: str = Form(...),  # Full JSON config for sheets/assets/columns
    existing_job_id: str = Form(...),  # Job ID of existing session
    background_tasks: BackgroundTasks = None
):
    """Add assets to an existing session"""
    
    # Verify the existing job exists
    if existing_job_id not in progress_map:
        return JSONResponse({"error": "Existing session not found"}, status_code=404)
    
    # Create a temporary session directory for processing
    session_dir = os.path.join(OUTPUT_BASE, f"temp_{existing_job_id}")
    os.makedirs(session_dir, exist_ok=True)
    
    parsed_config = json.loads(config)
    
    # Create a temporary file for the new upload
    input_path = os.path.join(f"additional_{file.filename}")
    with open(input_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # Run the processing job in the background, but append to existing images
    background_tasks.add_task(run_processing_job_add_assets, existing_job_id, input_path, parsed_config, session_dir)

    return JSONResponse({"job_id": existing_job_id})


async def run_processing_job_add_assets(job_id, input_path, parsed_config, session_dir):
    """Process additional assets and append to existing ones"""
    update_progress(job_id, "Processing additional assets", 5)
    wb = load_workbook(input_path)

    sheets_data = parsed_config["sheets"]

    # Keep only selected sheets
    selected_indices = [s["index"] for s in sheets_data]
    all_sheetnames = wb.sheetnames

    for idx, name in enumerate(all_sheetnames, start=1):
        if idx not in selected_indices:
            ws_to_remove = wb[name]
            wb.remove(ws_to_remove)

    for i, sheet_conf in enumerate(sheets_data):
        ws = wb.worksheets[sheet_conf["index"] - 1]
        ranges = [(a["start"], a["end"]) for a in sheet_conf["assets"]]
        columns_range = sheet_conf["columns"]
        process_sheet(ws, ranges, columns_range)
        update_progress(job_id, f"Processing additional sheet {sheet_conf['index']}", 10 + int((i + 1) / len(sheets_data) * 20))
        await asyncio.sleep(0.1)

    modified_excel = os.path.join(session_dir, "additional_processed.xlsx")
    wb.save(modified_excel)

    update_progress(job_id, "Converting additions to PDF", 40)
    subprocess.run(f'soffice --headless --convert-to pdf "{modified_excel}" --outdir "{session_dir}"', shell=True, check=True)
    await asyncio.sleep(0.1)

    update_progress(job_id, "Converting additions to PNG", 60)
    subprocess.run(f'pdftoppm -png -rx 300 -ry 300 "{os.path.join(session_dir,"additional_processed.pdf")}" "{os.path.join(session_dir,"additional_page")}"', shell=True, check=True)

    # Get existing images
    existing_images_temp = sorted([f for f in os.listdir("outputs") if f.endswith(".png")])
    existing_images = []
    for url in existing_images_temp:
        existing_images.append(f"/files/{url}")
    # Process new images
    new_png_urls = []
    all_files = sorted([f for f in os.listdir(session_dir) if f.endswith(".png")])
    for idx, f in enumerate(all_files):
        # Create unique filename for output folder
        unique_filename = f"{job_id}_additional_{f}"
        output_path = os.path.join(OUTPUT_BASE, unique_filename)
        
        # Crop and save directly to outputs folder
        crop_image(os.path.join(session_dir, f), output_path)
        update_progress(job_id, "Cropping additional images", 60 + int((idx + 1) / len(all_files) * 35))
        await asyncio.sleep(0.1)
        new_png_urls.append(f"/files/{unique_filename}")

    # Clean up the temporary session directory
    try:
        shutil.rmtree(session_dir, ignore_errors=True)
        print(f"Cleaned up temporary session directory: {session_dir}")
    except Exception as e:
        print(f"Error cleaning up session directory: {e}")

    # Combine existing and new images
    all_images = existing_images + new_png_urls
    
    update_progress(job_id, "Done", 100)
    progress_map[job_id]["images"] = all_images


@app.post("/add-pngs/")
async def add_pngs(
    files: list[UploadFile],
    existing_job_id: str = Form(...),  # Job ID of existing session
):
    """Add PNG files to an existing session"""
    
    # Verify the existing job exists
    if existing_job_id not in progress_map:
        return JSONResponse({"error": "Existing session not found"}, status_code=404)
    
    # Get existing images
    existing_images = progress_map[existing_job_id].get("images", [])
    
    # Process uploaded PNG files
    new_png_urls = []
    for i, file in enumerate(files):
        if file.content_type == "image/png" or file.filename.lower().endswith('.png'):
            # Create unique filename
            random_id = str(uuid4())
            unique_filename = f"{existing_job_id}_{random_id}.png"
            output_path = os.path.join(OUTPUT_BASE, unique_filename)
            
            # Save the PNG file directly
            with open(output_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            new_png_urls.append(f"/files/{unique_filename}")
    
    # Combine existing and new images
    all_images = existing_images + new_png_urls
    
    # Update progress map
    progress_map[existing_job_id]["images"] = all_images
    
    return JSONResponse({
        "job_id": existing_job_id,
        "added_count": len(new_png_urls),
        "total_images": len(all_images),
        "images": all_images
    })


@app.post("/inspect-excel/")
async def inspect_excel(file: UploadFile):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, read_only=True)
    sheets = [{"index": i + 1, "name": name} for i, name in enumerate(wb.sheetnames)]
    return {"sheets": sheets}
